"""Import în masă din Excel: șablon descărcabil + creare automată de dosare.

Fluxul: descărcați șablonul -> completați câte un rând per lucrător ->
încărcați fișierul -> aplicația creează dosarele și generează pachetele.
"""
import io
from datetime import date, datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill

# (antet Excel, cheie internă, exemplu, obligatoriu)
COLUMNS = [
    ("Nume *", "surname", "UDDIN", True),
    ("Prenume *", "given_name", "MD MAIN", True),
    ("Țara * (BANGLADESH/NEPAL)", "country", "BANGLADESH", True),
    ("Data nașterii * (ZZ.LL.AAAA)", "dob", "15.08.1988", True),
    ("Locul nașterii", "pob", "CUMILLA", False),
    ("Sex (M/F)", "sex", "M", False),
    ("Nr. pașaport *", "passport_number", "A19775898", True),
    ("Nr. personal", "personal_no", "4180794705", False),
    ("Data eliberării (ZZ.LL.AAAA)", "doi", "15.08.2025", False),
    ("Valabil până la * (ZZ.LL.AAAA)", "doe", "14.08.2035", True),
    ("Eliberat de", "authority", "DHAKA", False),
    ("Prenume tată (doar Bangladesh)", "father", "ABDUL MOMIN", False),
    ("Prenume mamă (doar Bangladesh)", "mother", "KOHINUR BEGUM", False),
    ("Adresă permanentă (doar Bangladesh)", "perm_addr",
     "FULMORI, CHAUDDAGRAM MESHTOLI BAZAR - 3550 CUMILLA", False),
    ("Funcția RO *", "job_title", "Curier livrator", True),
    ("Funcția EN (pt. CIM)", "job_title_en", "Delivery courier", False),
    ("Cod CORM", "occupation_code", "962101", False),
    ("Adresa cazare * (din Comodat)", "address",
     "mun. Chișinău, str. București, nr. 83, ap. 9", True),
    ("Tip contract (locațiune/comodat)", "contract_type", "locațiune", False),
    ("Nr. contract", "contract_number", "F/N", False),
    ("Data contract * (ZZ.LL.AAAA)", "contract_date", "31.10.2025", True),
    ("Nr. demers *", "demers_number", "8B/25", True),
    ("Data demers (ZZ.LL.AAAA)", "demers_date", "23.12.2025", False),
    ("Nr. CIM *", "cim_number", "8B/26", True),
    ("Data CIM (ZZ.LL.AAAA)", "cim_date", "19.03.2026", False),
]


def build_template() -> bytes:
    """Șablonul Excel: foaia 'Lucrători' (de completat) + foaia 'Exemplu'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Lucrători"
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="0D3B66")
    for col, (header, _, example, _req) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[cell.column_letter].width = max(16, min(40, len(header) + 4))

    ex = wb.create_sheet("Exemplu")
    for col, (header, _, example, _req) in enumerate(COLUMNS, start=1):
        c1 = ex.cell(row=1, column=col, value=header)
        c1.font = header_font
        c1.fill = header_fill
        ex.cell(row=2, column=col, value=example)
        ex.column_dimensions[c1.column_letter].width = max(16, min(40, len(header) + 4))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _parse_date(value, field: str):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    for fmt in ("%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    raise ValueError(f"{field}: dată invalidă „{s}” (folosiți ZZ.LL.AAAA)")


def parse_workbook(file_bytes: bytes) -> tuple[list[dict], list[str]]:
    """Citește foaia 'Lucrători'. Returnează (rânduri_valide, erori)."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = wb["Lucrători"] if "Lucrători" in wb.sheetnames else wb.active

    rows, errors = [], []
    for i, xl_row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not xl_row or all(v in (None, "") for v in xl_row):
            continue
        raw = {key: (xl_row[j] if j < len(xl_row) else None)
               for j, (_, key, _, _) in enumerate(COLUMNS)}
        try:
            row = _normalize_row(raw)
            rows.append(row)
        except ValueError as e:
            name = f"{raw.get('surname') or '?'} {raw.get('given_name') or ''}".strip()
            errors.append(f"Rândul {i} ({name}): {e}")
    return rows, errors


def _normalize_row(raw: dict) -> dict:
    row = {}
    for _, key, _, _ in COLUMNS:
        v = raw.get(key)
        row[key] = str(v).strip() if v not in (None, "") else ""

    # obligatorii
    missing = [h for h, k, _, req in COLUMNS if req and not row[k]]
    if missing:
        raise ValueError("lipsesc: " + ", ".join(m.replace(" *", "") for m in missing))

    # țara
    country = row["country"].upper().replace("★", "").strip()
    if country not in ("BANGLADESH", "NEPAL"):
        raise ValueError(f"țara „{row['country']}” nu e suportată (BANGLADESH/NEPAL)")
    row["country"] = country

    # date calendaristice
    for key, label in (("dob", "Data nașterii"), ("doi", "Data eliberării"),
                       ("doe", "Valabil până la"), ("contract_date", "Data contract"),
                       ("demers_date", "Data demers"), ("cim_date", "Data CIM")):
        row[key] = _parse_date(raw.get(key), label)

    # valori implicite
    row["sex"] = (row["sex"] or "M").upper()[:1]
    row["contract_type"] = row["contract_type"] or "locațiune"
    row["contract_number"] = row["contract_number"] or "F/N"
    if not row["authority"]:
        row["authority"] = "DHAKA" if country == "BANGLADESH" else "MOFA, DEPARTMENT OF PASSPORTS"
    row["demers_date"] = row["demers_date"] or date.today()
    row["cim_date"] = row["cim_date"] or date.today()
    row["doi"] = row["doi"] or date.today()

    # regula Bangladesh
    if country == "BANGLADESH" and not row["father"]:
        raise ValueError("Bangladesh: „Prenume tată” este obligatoriu (pagina 2 a pașaportului)")

    for k in ("surname", "given_name", "pob", "father", "mother", "perm_addr",
              "passport_number", "authority"):
        row[k] = row[k].upper() if isinstance(row[k], str) else row[k]
    return row


def row_to_package(row: dict):
    """Construiește WorkerPackage dintr-un rând normalizat."""
    from models import (Accommodation, Country, Employment, PackageMeta,
                        PassportData, WorkerPackage)
    return WorkerPackage(
        passport=PassportData(
            surname=row["surname"], given_name=row["given_name"],
            country=Country(row["country"]),
            date_of_birth=row["dob"], place_of_birth=row["pob"] or "-",
            sex=row["sex"], passport_number=row["passport_number"],
            personal_no=row["personal_no"] or None,
            date_of_issue=row["doi"], date_of_expiry=row["doe"],
            issuing_authority=row["authority"],
            father_name=row["father"] or None, mother_name=row["mother"] or None,
            permanent_address=row["perm_addr"] or None),
        accommodation=Accommodation(
            address=row["address"], contract_type=row["contract_type"],
            contract_number=row["contract_number"], contract_date=row["contract_date"]),
        employment=Employment(
            job_title=row["job_title"], job_title_en=row["job_title_en"],
            occupation_code=row["occupation_code"]),
        meta=PackageMeta(
            demers_number=row["demers_number"], demers_date=row["demers_date"],
            cim_number=row["cim_number"], cim_date=row["cim_date"]),
    )


def process_bulk(rows: list[dict]) -> tuple[list[dict], bytes]:
    """Creează dosare + pachete pentru toate rândurile.

    Returnează (rezultate, zip_combinat). Fiecare rezultat:
    {nume, status: 'ok'|'avertisment'|'eroare', detalii}
    """
    import zipfile

    from generators import PACKAGE_DOCUMENTS, generate_package_zip
    from utils import storage
    from utils.validators import check_package_consistency

    results = []
    combined = io.BytesIO()
    with zipfile.ZipFile(combined, "w", zipfile.ZIP_DEFLATED) as big:
        for row in rows:
            name = f"{row['surname']} {row['given_name']}".strip()
            try:
                pkg = row_to_package(row)
                storage.create_dosar(name, row["country"])
                zip_bytes, _files = generate_package_zip(pkg, PACKAGE_DOCUMENTS)
                files = {}
                with zipfile.ZipFile(io.BytesIO(zip_bytes)) as zf:
                    for n in zf.namelist():
                        files[n] = zf.read(n)
                        storage.save_generated(name, n, files[n])
                zip_name = f"PACHET_{storage._safe_name(name)}.zip"
                storage.save_generated(name, zip_name, zip_bytes)
                storage.update_meta(name, adresa_cazare=row["address"],
                                    status="acte generate — de verificat (import Excel)",
                                    generat_la=datetime.now().strftime("%d.%m.%Y %H:%M"))
                big.writestr(zip_name, zip_bytes)
                warns = check_package_consistency(files, pkg)
                if warns:
                    results.append({"nume": name, "status": "avertisment",
                                    "detalii": " | ".join(warns)})
                else:
                    results.append({"nume": name, "status": "ok",
                                    "detalii": f"{len(files)} documente generate"})
            except Exception as e:
                results.append({"nume": name, "status": "eroare", "detalii": str(e)})
    return results, combined.getvalue()
